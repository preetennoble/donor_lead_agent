from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from urllib.parse import urlparse

def generate_research_excel(company_data: dict) -> BytesIO:
    """
    Generate an Excel sheet containing company research data.
    Returns a BytesIO object with the Excel content.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Research"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    italic_font = Font(name="Calibri", size=11, italic=True, color="595959")
    
    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Blue
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # Steel Blue
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Soft Gray
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # 1. Title Row
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = f"Company Research Report: {company_data.get('company_name', 'Unknown')}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # 2. Metadata Info
    ws["A2"] = "Generated On:"
    ws["A2"].font = bold_font
    ws["B2"] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
    ws["B2"].font = regular_font
    ws.row_dimensions[2].height = 20
    
    # Add an empty row
    ws.row_dimensions[3].height = 15
    
    # 3. Headers
    ws["A4"] = "Field Name"
    ws["A4"].font = header_font
    ws["A4"].fill = header_fill
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws["B4"] = "Value / Research Data"
    ws["B4"].font = header_font
    ws["B4"].fill = header_fill
    ws["B4"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 25
    
    # Styles
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Soft Ice Blue
    section_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    
    # 4. Extract Data Sources
    research = company_data.get('research_json', {}) or {}
    csr_data = company_data.get('csr_data', {}) or {}
    fin_data = company_data.get('financial_data', {}) or {}
    edu_evidence = company_data.get('education_fitment_evidence', {}) or {}
    edu_spend = company_data.get('education_spend', {}) or {}
    csr_budget = company_data.get('csr_budget', {}) or {}
    
    partners = csr_data.get('implementation_partners') or research.get('existing_implementation_partners') or []
    committee = csr_data.get('committee_members') or company_data.get('committee_members') or []
    fys = fin_data.get('fiscal_years') or []
    
    sections = []
    
    # Section: Company Overview & General Research
    overview_fields = [
        ('Industry', research.get('industry', 'Not Found')),
        ('Website', research.get('website', 'Not Found')),
        ('City', research.get('city', 'Not Found')),
        ('State', research.get('state', 'Not Found')),
        ('Has Company Foundation', research.get('has_company_foundation', 'Not Found')),
        ('Data Confidence', research.get('confidence', 'unverified')),
        ('Source URL', research.get('source_url', 'Not Found')),
    ]
    sections.append(("1. COMPANY OVERVIEW", overview_fields))
    
    # Section: Comprehensive Financial Data (Multi-Year)
    fin_fields = []
    if fys:
        for fy in fys:
            to_val = (fin_data.get('turnover') or {}).get(fy)
            pbt_val = (fin_data.get('pbt') or {}).get(fy)
            np_val = (fin_data.get('net_profit') or {}).get(fy)
            nw_val = (fin_data.get('net_worth') or {}).get(fy)
            
            to_str = f"₹{to_val:,.2f} Cr" if to_val is not None else "-"
            pbt_str = f"₹{pbt_val:,.2f} Cr" if pbt_val is not None else "-"
            np_str = f"₹{np_val:,.2f} Cr" if np_val is not None else "-"
            nw_str = f"₹{nw_val:,.2f} Cr" if nw_val is not None else "-"
            
            fin_fields.append((f"{fy} Financials", f"Turnover: {to_str} | PBT: {pbt_str} | Net Profit: {np_str} | Net Worth: {nw_str}"))
    else:
        fin_fields.append(('Financial Data', 'Not Available'))
        
    if csr_budget.get('csr_budget_2pct') is not None:
        fin_fields.append(('CSR Budget (2% Avg PBT)', f"₹{csr_budget.get('csr_budget_2pct'):,.2f} Cr (Avg 3-yr PBT: ₹{csr_budget.get('average_pbt'):,.2f} Cr)"))
        
    sections.append(("2. FINANCIAL DATA & METRICS", fin_fields))
    
    # Section: 6 Education Fitment Checks
    education_check_defs = [
        ('csr_stem_education', 'STEM Education'),
        ('csr_school_infra_transformation', 'School Infrastructure Transformation'),
        ('csr_holistic_transformation', 'Holistic School Transformation'),
        ('csr_anganwadi_transformation', 'Anganwadi Transformation'),
        ('csr_quality_education', 'Quality Education'),
        ('csr_model_school_transformation', 'Model School Transformation')
    ]
    edu_fields = []
    for field_key, field_title in education_check_defs:
        val = research.get(field_key, 'Not Found')
        ev = edu_evidence.get(field_key, {}) or {}
        details = [f"Result: {val}"]
        if ev.get('status'):
            details.append(f"Status: {ev.get('status')}")
        if ev.get('evidence'):
            details.append(f"Evidence: {ev.get('evidence')}")
        if ev.get('sources'):
            srcs = [s if isinstance(s, str) else s.get('url', '') for s in ev.get('sources', [])]
            srcs = [s for s in srcs if s]
            if srcs:
                details.append(f"Sources: {', '.join(srcs)}")
        edu_fields.append((field_title, "\n".join(details)))
        
    sections.append(("3. 6 DEDICATED EDUCATION FITMENT CHECKS", edu_fields))
    
    # Section: CSR Profile & Governance
    csr_spend_val = csr_data.get('csr_spend')
    csr_spend_yr = csr_data.get('csr_spend_year')
    csr_unspent = csr_data.get('csr_unspent_amount')
    
    annual_csr_str = f"₹{csr_spend_val:,.2f} Cr ({csr_spend_yr})" if (csr_spend_val is not None and csr_spend_yr) else (f"₹{csr_spend_val:,.2f} Cr" if csr_spend_val is not None else research.get('education_csr_spend', 'Not Found'))
    unspent_str = f"₹{csr_unspent:,.2f} Cr" if csr_unspent is not None else "Not Found"
    
    # Format Committee Members with LinkedIn URLs
    comm_linkedin = company_data.get('committee_members_linkedin') or {}
    if committee:
        comm_strs = []
        for m in committee:
            li_url = comm_linkedin.get(m)
            if li_url:
                comm_strs.append(f"{m} (LinkedIn: {li_url})")
            else:
                comm_strs.append(m)
        committee_str = "\n".join(comm_strs) if len(comm_strs) > 1 else comm_strs[0]
    else:
        committee_str = "Not Found"

    csr_fields = [
        ('CSR Themes', ', '.join(research.get('thematic_focus', [])) if research.get('thematic_focus') else 'Not Found'),
        ('Geography', research.get('geographical_priority', 'Not Found')),
        ('CSR Focus', research.get('company_csr_focus', 'Not Found')),
        ('Past CSR Programs', research.get('previous_education_projects', 'Not Found')),
        ('Avg Ticket Size', research.get('avg_ticket_size', 'Not Found')),
        ('Program District/State', research.get('program_district_state', 'Not Found')),
        ('Annual CSR Spend', annual_csr_str),
        ('CSR Unspent Amount', unspent_str),
        ('Education CSR Spend', f"₹{edu_spend.get('current_education_spend'):,.2f} Cr ({edu_spend.get('current_education_percentage')}% of CSR)" if edu_spend.get('current_education_spend') is not None else research.get('education_csr_spend', 'Not Found')),
        ('CSR Spend (Previous FY)', research.get('csr_spend_previous_fy', 'Not Found')),
        ('CSR Spend (3 FY)', research.get('csr_spend_previous_3fy', 'Not Found')),
        ('Implementation Partners (NGOs)', ', '.join(partners) if partners else 'Not Found'),
        ('CSR Committee Members', committee_str),
    ]
    sections.append(("4. CSR STRATEGY & GOVERNANCE", csr_fields))

    current_row = 5
    for sec_title, fields in sections:
        # Section Header Row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        sec_cell = ws.cell(row=current_row, column=1, value=sec_title)
        sec_cell.font = section_font
        sec_cell.fill = section_fill
        sec_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 28
        current_row += 1
        
        # Section Rows
        for field_name, field_value in fields:
            cell_a = ws.cell(row=current_row, column=1, value=field_name)
            cell_b = ws.cell(row=current_row, column=2)
            
            cell_a.font = bold_font
            cell_a.border = thin_border
            cell_b.border = thin_border
            
            if current_row % 2 == 1:
                cell_a.fill = zebra_fill
                cell_b.fill = zebra_fill
                
            if field_value == 'Not Found' or not field_value or field_value == '-':
                cell_b.value = "Not Found"
                cell_b.font = italic_font
            else:
                cell_b.value = str(field_value)
                cell_b.font = regular_font
                
            cell_a.alignment = Alignment(vertical="top")
            cell_b.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Dynamic row height based on line breaks
            num_lines = str(field_value).count('\n') + 1
            ws.row_dimensions[current_row].height = max(20, num_lines * 18)
            current_row += 1
            
        # Spacer row between sections
        ws.row_dimensions[current_row].height = 10
        current_row += 1
        
    # Auto-fit column widths
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 80
    
    # Save spreadsheet to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_research_excel_filename(company_name: str) -> str:
    """Generate a clean filename for the Excel sheet."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    return f"Research_{safe_name}_{timestamp}.xlsx"


def generate_combined_research_excel(companies: list) -> BytesIO:
    """Generate the approved, fixed-column company comparison worksheet."""
    columns = [
        "company name", "industry", "city", "state", "csr theme", "geography",
        "csr focus", "past csr program", "Program District/State", "source url",
        "Tier ", "Company Score", "ennoble fitment", "STEM Education", "School Infra Transformation",
        "Holistic School Transformation", "Anganwadi Transformation", "Quality Education",
        "Model School Transformation", "CSR Contact Person name",
        "CSR Contact Person designation", "financial data annual turnover Annual Turnover (₹ Cr)",
        "PBT (₹ Cr)", "Net Profit (₹ Cr)", "Net Worth (₹ Cr)",
        "CSR Budget Calculation (Average PBT basis)",
        "CSR Budget Calculation (Net Profit basis)",
        "Annual CSR Spend ", "CSR Unspent Amount", "csr comittie member", "implementation ngo partner ",
        "LinkedIn",
    ]

    def display(value):
        if value is None or value == "" or value == [] or value == {}:
            return "Not Found"
        if isinstance(value, (list, tuple)):
            return ", ".join(str(item) for item in value) or "Not Found"
        return value

    def education_fit(value):
        """Normalize education-fit answers to the report's four categories."""
        if value is None or str(value).strip() == "":
            return "Not Evident"
        normalized = str(value).strip().casefold()
        if normalized in {"yes", "high", "high fit"}:
            return "High"
        if normalized in {"medium", "medium fit", "moderate"}:
            return "Medium"
        if normalized in {"no", "low", "low fit"}:
            return "Low"
        if normalized in {"not evident", "not found", "unknown", "n/a", "na"}:
            return "Not Evident"
        return str(value)

    def latest_financial(financial_data, field):
        years = financial_data.get("fiscal_years") or []
        values = financial_data.get(field) or {}
        for year in years:
            if values.get(year) is not None:
                return values[year]
        return next((value for value in values.values() if value is not None), "Not Found")

    def financial_block(financial_data, field):
        values = financial_data.get(field) or {}
        years = financial_data.get("fiscal_years") or list(values.keys())
        lines = [f"{year}: {values.get(year)}" for year in years if values.get(year) is not None]
        return "\n".join(lines) or "Not Found"

    def source_formula(research):
        raw = research.get("source_url") or ""
        urls = [url.strip() for url in str(raw).split(";") if url.strip() and url.strip().startswith("http")]
        formulas = []
        for url in urls:
            label = urlparse(url).netloc or url
            formulas.append(f'HYPERLINK("{url.replace(chr(34), chr(34) * 2)}","{label.replace(chr(34), chr(34) * 2)}")')
        return "=" + "&CHAR(10)&".join(formulas) if formulas else display(raw)

    def committee_formula(company):
        csr = company.get("csr_data") or {}
        members = csr.get("committee_members") or company.get("committee_members") or []
        links = company.get("committee_members_linkedin") or csr.get("committee_members_linkedin") or {}
        formulas = []
        plain_names = []
        for member in members:
            name = str(member).strip()
            url = links.get(member) or links.get(name)
            if not url:
                for link_name, link_url in links.items():
                    if str(link_name).strip().casefold() == name.casefold():
                        url = link_url
                        break
            if url:
                safe_url = str(url).replace('"', '""')
                safe_name = name.replace('"', '""')
                formulas.append(f'HYPERLINK("{safe_url}","{safe_name}")')
            else:
                plain_names.append(name)
        if formulas:
            formula = "=" + "&CHAR(10)&".join(formulas)
            if plain_names:
                formula += '&CHAR(10)&"' + '"&CHAR(10)&"'.join(n.replace('"', '""') for n in plain_names) + '"'
            return formula
        return "\n".join(plain_names) or "Not Found"

    rows = []
    for company in companies:
        research = company.get("research_json") or {}
        csr = company.get("csr_data") or {}
        financial = company.get("financial_data") or {}
        contact = research.get("contact") or {}
        fitment = company.get("program_fitment") or {}
        row = {
            "company name": display(company.get("company_name")),
            "industry": display(research.get("industry")),
            "city": display(research.get("city")),
            "state": display(research.get("state")),
            "csr theme": display(research.get("thematic_focus")),
            "geography": display(research.get("geographical_priority")),
            "csr focus": display(research.get("company_csr_focus")),
            "past csr program": display(research.get("previous_education_projects")),
            "Program District/State": display(research.get("program_district_state")),
            "source url": source_formula(research),
            "Tier ": display(company.get("tier") or company.get("category")),
            "Company Score": display(company.get("score")),
            "ennoble fitment": display(fitment.get("Ennoble Fitment")),
            "STEM Education": education_fit(research.get("csr_stem_education")),
            "School Infra Transformation": education_fit(research.get("csr_school_infra_transformation")),
            "Holistic School Transformation": education_fit(research.get("csr_holistic_transformation")),
            "Anganwadi Transformation": education_fit(research.get("csr_anganwadi_transformation")),
            "Quality Education": education_fit(research.get("csr_quality_education")),
            "Model School Transformation": education_fit(research.get("csr_model_school_transformation")),
            "CSR Contact Person name": display(" ".join(str(v).strip() for v in [contact.get("first_name", ""), contact.get("last_name", "")] if v and str(v).strip())),
            "CSR Contact Person designation": display(contact.get("designation")),
            "financial data annual turnover Annual Turnover (₹ Cr)": financial_block(financial, "turnover"),
            "PBT (₹ Cr)": financial_block(financial, "pbt"),
            "Net Profit (₹ Cr)": financial_block(financial, "net_profit"),
            "Net Worth (₹ Cr)": financial_block(financial, "net_worth"),
            "CSR Budget Calculation (Average PBT basis)": display((company.get("csr_budget") or {}).get("csr_budget_2pct")),
            "CSR Budget Calculation (Net Profit basis)": display((company.get("csr_budget_net_profit") or {}).get("csr_budget_2pct_net_profit")),
            "Annual CSR Spend ": display(csr.get("csr_spend")),
            "CSR Unspent Amount": display(csr.get("csr_unspent_amount")),
            "csr comittie member": committee_formula(company),
            "implementation ngo partner ": display(csr.get("implementation_partners") or research.get("existing_implementation_partners")),
            "LinkedIn": display(contact.get("linkedin") or contact.get("linkedin_url")),
        }
        rows.append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "Selected Companies"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1" if columns else "A1"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_index, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_index, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, 2):
        for col_index, column in enumerate(columns, 1):
            cell = ws.cell(row=row_index, column=col_index, value=row.get(column, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column == "LinkedIn" and str(row.get(column, "")).startswith("http"):
                cell.hyperlink = row[column]
                cell.font = Font(color="0563C1", underline="single")
            elif column in {"source url", "csr comittie member"} and str(row.get(column, "")).startswith("="):
                cell.font = Font(color="0563C1", underline="single")

    for col_index, column in enumerate(columns, 1):
        values = [str(ws.cell(row=r, column=col_index).value or "") for r in range(1, min(ws.max_row, 20) + 1)]
        ws.column_dimensions[ws.cell(row=1, column=col_index).column_letter].width = min(max(max(map(len, values), default=10) + 2, 14), 42)
    ws.row_dimensions[1].height = 32

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_combined_research_excel_filename() -> str:
    return f"Selected_Company_Research_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
